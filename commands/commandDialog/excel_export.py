"""Cut-list Excel export.

Walks every cabinet in the active design and lists only the boards that are
actually *visible* (a hidden component -- or a hidden container like the
``ukrute``/``Ultrabox`` wrapper -- is left out, matching what the user sees).
Each visible board's DULJINA/ŠIRINA are **measured from the real 3D geometry**
(not the parametric ``size`` expressions), so any manual edit a user makes to a
board after the add-in creates it is reflected; the banding is reduced to the
(long, short) counts the woodshop needs.

**One workbook per (decor, thickness) combination.**  A design can mix several
board decors (colours) and several panel thicknesses (typically the 18mm base
material and a thin back-panel material), and each is a physically distinct
sheet good the woodshop orders separately -- so each combination gets its own
copy of the Elgrad supplier order form (``narudzba-excel.xlsm``), with the
decor name written into the sheet's IVERAL / RUB. TRAKA DEKOR cells and
prepended to the output filename along with the thickness.

Dimensions (and the thickness used to bucket boards into workbooks) come from
the greatest-area-faces method (see ``_board_geometry_dims``): a board's two
broad faces are its largest, their normal is the thickness axis, and
length/width are the in-plane extent of the body's vertices -- measured in the
board's *own* frame, so it stays correct even for a door swung open on its
joint, and robust to grooves/split cuts (which only cut inward).

No adsk model editing happens here: we only *read* geometry and finish state,
then hand pure data rows to openpyxl.  That keeps the export independent of the
command-dialog session and its preview/rollback lifecycle.
"""

import math
import os
import re

import adsk.core
import adsk.fusion

from . import base_design
from ...lib import fusionAddInUtils as futil


# ---------------------------------------------------------------------------
# Elementi sheet layout (1-based columns).  Rows 1-3 are the material/header
# block the user maintains by hand; data starts at row 4.  Column A (R.B.) is
# pre-numbered 1..99 in the template, so we only write B..I and leave A alone.
# ---------------------------------------------------------------------------
_SHEET_NAME = "Elementi"
_FIRST_DATA_ROW = 4
_LAST_TEMPLATE_ROW = 102  # rows 4..102 are the pre-formatted data band

_COL_RB = 1          # A  R.B.            (pre-filled, left intact)
_COL_DULJINA = 2     # B  DULJINA         longer face dimension, cm
_COL_SIRINA = 3      # C  ŠIRINA          shorter face dimension, cm
_COL_KOM = 4         # D  BR.KOM.         quantity
_COL_MEL_LONG = 5    # E  mini/mel ABS long  (unused -> blank; single band type)
_COL_MEL_SHORT = 6   # F  mini/mel ABS short (unused -> blank)
_COL_ABS_LONG = 7    # G  ABS 2mm long    banded long-edge count
_COL_ABS_SHORT = 8   # H  ABS 2mm short   banded short-edge count
_COL_NAPOMENA = 9    # I  NAPOMENA        board name (Croatian)

# Header block (row 1): the board material ("IVERAL", merged B1:D1) and edge
# band ("RUB. TRAKA DEKOR", merged G1:H1) decor-name cells.  One workbook is
# written per (decor, thickness) bucket, so both get the bucket's decor name.
_ROW_HEADER = 1
_COL_IVERAL = 2      # B1  IVERAL           board decor name
_COL_RUB_TRAKA = 7   # G1  RUB. TRAKA DEKOR edge band decor name


# Row order in the sheet: bottom -> sides -> back -> top -> divider -> shelves
# -> plinth -> doors -> stiffeners -> ultrabox (podnica/zadnja/fronta).  Any
# board not listed here sorts last, in name order.
_ROW_ORDER = [
    "donja ploca", "bok desno", "bok lijevo", "ledja", "gornja ploca",
    "pregrada", "polica", "cokla", "fronta desno", "fronta lijevo",
    "ukruta otraga", "podnica", "zadnja", "fronta",
]

# NAPOMENA (col I) text per board; ultrabox parts are disambiguated from the
# cabinet's own doors ("fronta desno"/"fronta lijevo" vs the ultrabox "fronta").
_NAPOMENA = {
    "podnica": "ultrabox podnica",
    "zadnja": "ultrabox zadnja",
    "fronta": "ultrabox fronta",
}


# ---------------------------------------------------------------------------
# Data gathering (pure reads + expression evaluation, no model edits).
# ---------------------------------------------------------------------------
def _cabinet_flags(design, prefix):
    """This cabinet's parameter values keyed by base name (prefix stripped)."""
    flags = {}
    for i in range(design.userParameters.count):
        p = design.userParameters.item(i)
        if p.name.startswith(prefix):
            flags[p.name[len(prefix):]] = p.value
    return flags


def _get_prefixes(design):
    """Every cabinet prefix in the design (e.g. 'J1_', 'O1_'), discovered by
    intersecting which prefixes exist for every base_design user-parameter name
    -- independent of the dialog, so export works even with no dialog open."""
    param_names = {design.userParameters.item(i).name
                   for i in range(design.userParameters.count)}
    prefixes = None
    for raw_name, _expr, _unit in base_design.USER_PARAMS:
        base = raw_name.replace("{p}", "")
        matches = {n[: -len(base)] for n in param_names
                   if n.endswith(base) and len(n) > len(base)}
        prefixes = matches if prefixes is None else (prefixes & matches)
        if not prefixes:
            return []
    return sorted(prefixes or [])


def _extent(points, axis):
    """Span of a point cloud projected onto a (unit) axis vector."""
    vals = [p[0] * axis[0] + p[1] * axis[1] + p[2] * axis[2] for p in points]
    return (max(vals) - min(vals)) if vals else 0.0


def _normalize(v):
    m = math.sqrt(v[0] * v[0] + v[1] * v[1] + v[2] * v[2])
    return (v[0] / m, v[1] / m, v[2] / m) if m else v


def _cross(a, b):
    return (a[1] * b[2] - a[2] * b[1],
            a[2] * b[0] - a[0] * b[2],
            a[0] * b[1] - a[1] * b[0])


def _dims_from_geometry(points, thickness_normal, inplane_dir):
    """Board dimensions measured *in the board's own frame*, so they are correct
    regardless of how the board is oriented in space (a door swung open on its
    joint, or a user-rotated part) and robust to grooves/cuts (which only cut
    inward, never past the panel extent).

    Given the point cloud of the board's vertices, the thickness axis (the
    largest face's normal) and one in-plane axis, returns
    ``(duljina, sirina, debljina, world_dims)``:

    * ``duljina`` / ``sirina`` -- the longer / shorter of the two in-plane
      extents (DULJINA / ŠIRINA); the thickness extent is excluded.
    * ``debljina`` -- the panel thickness (extent along the normal axis), used
      to bucket boards into a workbook per material thickness.
    * ``world_dims`` -- a ``(dx, dy, dz)`` tuple with each of the board's three
      own-axis extents assigned to the world axis it most aligns with, for the
      banding long/short reduction (which is keyed to the board spec's plane).
    """
    n = _normalize(thickness_normal)
    d1 = _normalize(inplane_dir)
    d2 = _normalize(_cross(n, d1))          # the other in-plane axis

    e1, e2, et = _extent(points, d1), _extent(points, d2), _extent(points, n)
    duljina, sirina = (e1, e2) if e1 >= e2 else (e2, e1)

    world = [0.0, 0.0, 0.0]
    for vec, ext in ((d1, e1), (d2, e2), (n, et)):
        world[max(range(3), key=lambda i: abs(vec[i]))] = ext
    return duljina, sirina, et, tuple(world)


def _board_geometry_dims(body):
    """Measure a solid board body from its actual geometry via the greatest-area
    faces: the two broad faces (largest area) define the thickness normal, and
    the board's length/width come from the extent of its vertices in-plane.
    Returns ``(duljina, sirina, debljina, world_dims)`` in cm, or ``None`` if the
    body has no usable planar geometry."""
    faces = body.faces
    if faces.count == 0:
        return None
    largest = max((faces.item(i) for i in range(faces.count)), key=lambda f: f.area)

    plane = adsk.core.Plane.cast(largest.geometry)
    if plane is None:                       # not a planar broad face -> can't classify
        return None
    normal = (plane.normal.x, plane.normal.y, plane.normal.z)

    # longest straight edge of the broad face gives one in-plane direction
    best_dir, best_len = None, -1.0
    edges = largest.edges
    for i in range(edges.count):
        edge = edges.item(i)
        sv, ev = edge.startVertex, edge.endVertex
        if sv is None or ev is None:
            continue
        a, b = sv.geometry, ev.geometry
        v = (b.x - a.x, b.y - a.y, b.z - a.z)
        length = v[0] * v[0] + v[1] * v[1] + v[2] * v[2]
        if length > best_len:
            best_len, best_dir = length, v
    if best_dir is None:
        return None

    verts = body.vertices
    points = [(verts.item(i).geometry.x,
               verts.item(i).geometry.y,
               verts.item(i).geometry.z) for i in range(verts.count)]
    return _dims_from_geometry(points, normal, best_dir)


def _row(cabinet, spec_name, duljina, sirina, world_dims, banded_orients, qty):
    """Build one cut-list row dict from a measured board.  ``duljina``/``sirina``
    and ``world_dims`` come from the real geometry; ``banded_orients`` is the set
    of edge orientations actually banded (from the same effective banding the
    model is painted with, so user overrides are honoured); ``cabinet`` is the
    cabinet name, prefixed onto the NAPOMENA description."""
    long_band, short_band = base_design.banding_counts_for_orientations(
        spec_name, banded_orients, world_dims
    )
    label = _NAPOMENA.get(spec_name, spec_name)
    return {
        "duljina": round(duljina, 1),
        "sirina": round(sirina, 1),
        "kom": qty,
        "abs_long": long_band,
        "abs_short": short_band,
        "napomena": f"{cabinet} {label}" if cabinet else label,
    }


def _collect_boards(occ, out, prefix=""):
    """Collect ``(spec_name, body)`` for every *visible* board in `occ`'s subtree.

    Respects hidden state exactly as the add-in sets it: an occurrence whose
    light bulb is off hides its whole subtree (this is how the `ukrute` and
    `Ultrabox` wrappers, and flag-driven panel hides, work), so a hidden board
    -- or hidden container -- contributes nothing.  Each solid body is one
    physical piece, so a shelf split in two by a divider yields two pieces, and
    each materialized `Ultrabox N` drawer contributes its child boards while the
    hidden `Ultrabox` template contributes nothing.
    """
    if not occ.isLightBulbOn:
        return
    spec_name = base_design.spec_name_for_component(occ.component.name, prefix)
    if spec_name is not None:
        for body in occ.bRepBodies:
            if body.isSolid:
                out.append((spec_name, body))
    for child in occ.childOccurrences:
        _collect_boards(child, out, prefix)


def _thickness_mm(debljina_cm):
    """Bucket a measured thickness (cm) to the nearest whole millimetre, so
    geometry noise (e.g. 1.7998 cm) doesn't split one physical thickness into
    several buckets."""
    return round(debljina_cm * 10)


def _gather_rows(design):
    """All cut-list rows for the design, bucketed by ``(decor, thickness_mm)``
    -- each bucket is one physical sheet good and becomes its own workbook --
    then grouped by cabinet then panel type within a bucket.  Only *visible*
    boards are listed, and every dimension (including the thickness used for
    bucketing) is measured from the actual 3D geometry, so a user's manual
    edits after cabinet creation are reflected; identical pieces collapse into
    one row with a quantity.

    Returns ``{(decor, thickness_mm): [row, ...]}``.
    """
    # Imported lazily: apply the model's *effective* decor/banding (rule
    # defaults + the user's manual overrides) so the cut list matches what the
    # model is painted with, instead of re-deriving them from rules alone.
    from . import utils

    buckets = {}
    for prefix in _get_prefixes(design):
        flags = _cabinet_flags(design, prefix)
        cabinet = prefix.rstrip("_")

        # Walk this cabinet's occurrence tree for visible board bodies.  Prefer
        # the named wrapper occurrence; fall back to the root occurrences for a
        # legacy no-wrapper (single-cabinet) design, matching apply_finish.
        boards = []
        wrapper = next((o for o in design.rootComponent.occurrences
                        if o.component.name == cabinet), None)
        if wrapper is not None:
            _collect_boards(wrapper, boards, prefix)
        else:
            for occ in design.rootComponent.occurrences:
                _collect_boards(occ, boards, prefix)
        holder = wrapper if wrapper is not None else design.rootComponent

        # decor / banded orientations are per board spec (not per instance) -> cache
        decor_cache = {}

        def _decor(spec_name):
            if spec_name not in decor_cache:
                decor_cache[spec_name] = utils.effective_decor(
                    design, prefix, spec_name, holder, flags
                )
            return decor_cache[spec_name]

        banded_cache = {}

        def _banded(spec_name):
            if spec_name not in banded_cache:
                banded_cache[spec_name] = tuple(
                    utils.effective_banding(design, prefix, spec_name, holder, flags)
                )
            return banded_cache[spec_name]

        # group identical measured pieces per bucket:
        # (decor, thickness_mm, spec, DULJINA, ŠIRINA) -> row data
        groups = {}
        for spec_name, body in boards:
            measured = _board_geometry_dims(body)
            if measured is None:
                futil.log(f"Cut list: could not measure a '{prefix}{spec_name}' body",
                          adsk.core.LogLevels.WarningLogLevel)
                continue
            duljina, sirina, debljina, world_dims = measured
            decor = _decor(spec_name)
            thickness_mm = _thickness_mm(debljina)
            key = (decor, thickness_mm, spec_name, round(duljina, 1), round(sirina, 1))
            if key in groups:
                groups[key]["qty"] += 1
            else:
                groups[key] = {"spec": spec_name, "duljina": duljina,
                               "sirina": sirina, "world": world_dims, "qty": 1}

        def _order(name):
            return _ROW_ORDER.index(name) if name in _ROW_ORDER else len(_ROW_ORDER)

        for key in sorted(groups, key=lambda k: (k[0], k[1], _order(k[2]), -k[3], -k[4])):
            decor, thickness_mm, spec_name, _duljina, _sirina = key
            g = groups[key]
            row = _row(cabinet, g["spec"], g["duljina"], g["sirina"],
                       g["world"], _banded(g["spec"]), g["qty"])
            buckets.setdefault((decor, thickness_mm), []).append(row)
    return buckets


# ---------------------------------------------------------------------------
# Excel writing.
# ---------------------------------------------------------------------------
def _template_path():
    # excel_export.py -> commandDialog -> commands -> repo root
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(root, "narudzba-excel.xlsm")


def _safe_name(name):
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip()


# ---------------------------------------------------------------------------
# Export folder preference.  Stored in a small JSON file in the user's home
# directory (not in the repo -- this is a per-machine preference, not project
# config) so a chosen export location survives across Fusion sessions and
# different designs; shared with iverpan_export.py since both exports save to
# the same user-chosen folder.
# ---------------------------------------------------------------------------
_PREFS_PATH = os.path.join(os.path.expanduser("~"), ".ormarparametric", "export_prefs.json")


def _load_prefs():
    import json
    try:
        with open(_PREFS_PATH, encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, ValueError):
        return {}


def _save_prefs(prefs):
    import json
    os.makedirs(os.path.dirname(_PREFS_PATH), exist_ok=True)
    with open(_PREFS_PATH, "w", encoding="utf-8") as fh:
        json.dump(prefs, fh, indent=2)


def get_export_folder():
    """The user's saved cut-list export folder, or ``None`` if never set or no
    longer present on disk (e.g. a removed external drive) -- callers fall
    back to Desktop/home in that case."""
    folder = _load_prefs().get("export_folder")
    return folder if folder and os.path.isdir(folder) else None


def set_export_folder(folder):
    """Save `folder` as the export location for future exports."""
    prefs = _load_prefs()
    prefs["export_folder"] = folder
    _save_prefs(prefs)


def _default_export_folder():
    """The user's saved export folder if set, else the Desktop (or home if
    there's no Desktop directory)."""
    folder = get_export_folder()
    if folder is not None:
        return folder
    home = os.path.expanduser("~")
    desktop = os.path.join(home, "Desktop")
    return desktop if os.path.isdir(desktop) else home


def unique_path(path):
    """If `path` already exists, append an auto-incrementing ' (N)' suffix
    before the extension until a free name is found."""
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{root} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def resolve_output_paths(paths):
    """Decide the final path for each of `paths`, asking the user once if any
    of them already exist: "Yes" overwrites the colliding names as-is, "No"
    (or closing the dialog) auto-suffixes just the colliding ones via
    `unique_path` so nothing already on disk is lost.  Non-colliding paths are
    always returned unchanged.  Shared with iverpan_export.py since both write
    straight to a fixed folder with no save dialog of their own."""
    existing = [p for p in paths if os.path.exists(p)]
    if not existing:
        return list(paths)
    app = adsk.core.Application.get()
    ui = app.userInterface
    names = "\n".join(os.path.basename(p) for p in existing)
    result = ui.messageBox(
        f"Sljedeće datoteke već postoje:\n{names}\n\n"
        "Prepisati ih? (Ne = spremi pod novim imenom)",
        "Datoteka već postoji",
        adsk.core.MessageBoxButtonTypes.YesNoButtonType,
        adsk.core.MessageBoxIconTypes.QuestionIconType,
    )
    if result == adsk.core.DialogResults.DialogYes:
        return list(paths)
    return [unique_path(p) if p in existing else p for p in paths]


def _default_output_path(design, decor, thickness_mm):
    """Where to drop the exported copy: the user's saved export folder (see
    `get_export_folder`), else the Desktop, else the home directory.  The decor
    name and thickness are prepended so the per-(decor, thickness) workbooks
    for one design don't collide, and are identifiable to the woodshop at a
    glance."""
    app = adsk.core.Application.get()
    doc_name = _safe_name(app.activeDocument.name or "ormar") or "ormar"
    decor_name = _safe_name(decor) or "decor"
    folder = _default_export_folder()
    return os.path.join(folder, f"{decor_name}-{thickness_mm}mm-{doc_name}-krojna-lista.xlsm")


def _write_to_excel(rows, template_path, output_path, decor):
    """Load the template (preserving VBA + every other sheet), stamp the decor
    name into the IVERAL / RUB. TRAKA DEKOR header cells, clear the old data
    band, write the rows, and save a copy to output_path.  The template file is
    never modified."""
    import openpyxl

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb[_SHEET_NAME]

    ws.cell(row=_ROW_HEADER, column=_COL_IVERAL).value = decor
    ws.cell(row=_ROW_HEADER, column=_COL_RUB_TRAKA).value = decor

    # clear the pre-formatted data band (B..I), leaving column A's R.B. numbering
    for r in range(_FIRST_DATA_ROW, _LAST_TEMPLATE_ROW + 1):
        for c in range(_COL_DULJINA, _COL_NAPOMENA + 1):
            ws.cell(row=r, column=c).value = None

    for i, row in enumerate(rows):
        r = _FIRST_DATA_ROW + i
        ws.cell(row=r, column=_COL_RB).value = i + 1  # numbers rows past the band too
        ws.cell(row=r, column=_COL_DULJINA).value = row["duljina"]
        ws.cell(row=r, column=_COL_SIRINA).value = row["sirina"]
        ws.cell(row=r, column=_COL_KOM).value = row["kom"]
        # mini/mel ABS columns unused (single banding type) -> left blank
        # ABS 2mm banded-edge counts; blank rather than 0 to match the form's
        # own PRIMJER convention (zeros read as empty).
        ws.cell(row=r, column=_COL_ABS_LONG).value = row["abs_long"] or None
        ws.cell(row=r, column=_COL_ABS_SHORT).value = row["abs_short"] or None
        ws.cell(row=r, column=_COL_NAPOMENA).value = row["napomena"]

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Entry point (called from the dialog's export button).
# ---------------------------------------------------------------------------
def export_cut_list():
    """Export the active design's cut list to one .xlsm copy of the order form
    per (decor, thickness) combination present in the design -- each is a
    physically distinct sheet good the woodshop orders separately.

    Returns the list of output paths on success, or ``None`` if the design has
    no cabinets (nothing to export).  If a target filename already exists, the
    user is asked once (via `resolve_output_paths`) whether to overwrite it or
    save the colliding file(s) under a new ``' (N)'``-suffixed name.  Raises on
    I/O / openpyxl errors so the caller can surface them.
    """
    app = adsk.core.Application.get()
    design = adsk.fusion.Design.cast(app.activeProduct)
    if design is None:
        return None

    buckets = _gather_rows(design)
    if not buckets:
        return None

    template_path = _template_path()
    keys = sorted(buckets)
    planned_paths = resolve_output_paths(
        [_default_output_path(design, decor, thickness_mm) for decor, thickness_mm in keys]
    )

    output_paths = []
    for (decor, thickness_mm), output_path in zip(keys, planned_paths):
        rows = buckets[(decor, thickness_mm)]
        _write_to_excel(rows, template_path, output_path, decor)
        futil.log(f"Cut list exported: {len(rows)} rows -> {output_path}")
        output_paths.append(output_path)
    return output_paths


# ---------------------------------------------------------------------------
# Shared with iverpan_export.py: board discovery and geometry measurement are
# supplier-agnostic, so the Iverpan (single-sheet, one-row-per-piece) export
# reuses these rather than re-walking the occurrence tree a second way.
# ---------------------------------------------------------------------------
get_prefixes = _get_prefixes
cabinet_flags = _cabinet_flags
collect_boards = _collect_boards
board_geometry_dims = _board_geometry_dims
thickness_mm = _thickness_mm
row_order = _ROW_ORDER
napomena_by_spec = _NAPOMENA
