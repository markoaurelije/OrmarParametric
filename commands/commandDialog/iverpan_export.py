"""Cut-list export in the Iverpan supplier's order format
("Iverpan tablica za narudzbu.xlsx").

Unlike the Elgrad export (excel_export.py), Iverpan's ``Narudžba`` sheet carries
the material code and thickness *per row* (columns B/C), so the whole design's
cut list -- every cabinet, every decor, every thickness -- fits in one sheet;
there is no need to split into one workbook per decor/thickness.

**We do not know the user's real Iverpan catalogue codes.**  Iverpan identifies
board decors and edge-band tape by their own ``Šifra materijala`` / band-code
scheme (see the workbook's ``Dekori`` sheet), which has no relation to this
project's own decor names (``decors.json``).  Guessing a code would risk the
woodshop ordering the wrong board, so every material/band cell here is filled
with our own decor name as a plain-text placeholder -- the user is expected to
replace each with the real Iverpan code before sending the order.

Board discovery and geometry measurement are shared with excel_export.py (see
its "Shared with iverpan_export.py" section) since they're supplier-agnostic;
only the sheet layout, unit (mm here vs cm), and one-sheet-no-splitting
structure differ.
"""

import os
import re

import adsk.core
import adsk.fusion

from . import base_design
from . import excel_export as xls
from ...lib import fusionAddInUtils as futil


# ---------------------------------------------------------------------------
# Narudžba sheet layout (1-based columns).  Rows 1-13 are the header block
# (recipient info + column titles) the user maintains by hand; data starts at
# row 14.  Column A (R.b) is pre-numbered 1..792 in the template, so we only
# write B..K and leave A alone.
# ---------------------------------------------------------------------------
_SHEET_NAME = "Narudžba"
_FIRST_DATA_ROW = 14
_LAST_TEMPLATE_ROW = 805  # rows 14..805 are the pre-formatted data band

_COL_RB = 1            # A  R.b            (pre-filled, left intact)
_COL_MATERIJAL = 2     # B  Šifra materijala  -> our decor name (placeholder)
_COL_DEBLJINA = 3      # C  Deb.              board thickness, mm
_COL_MJERA1 = 4        # D  1.Mjera           longer face dimension, mm
_COL_MJERA2 = 5        # E  2. Mjera          shorter face dimension, mm
_COL_KOM = 6           # F  Br.               quantity
_COL_KANT_1A = 7       # G  kantiranje 1.mjera, edge A -> band decor name
_COL_KANT_1B = 8       # H  kantiranje 1.mjera, edge B -> band decor name
_COL_KANT_2A = 9       # I  kantiranje 2.mjera, edge A -> band decor name
_COL_KANT_2B = 10      # J  kantiranje 2.mjera, edge B -> band decor name
_COL_NAPOMENA = 11     # K  Napomena          "<cabinet> <board>"
_LAST_CLEAR_COL = 15   # O  last of the five Napomena columns


# ---------------------------------------------------------------------------
# Data gathering (pure reads, no model edits) -- mirrors excel_export._gather_rows
# but keeps every (decor, thickness) combination in one flat, ordered list
# instead of bucketing into separate workbooks, and measures in mm (Iverpan's
# own unit) rather than cm.
# ---------------------------------------------------------------------------
def _gather_rows(design):
    """All cut-list rows for the design, in one list ordered by cabinet then
    panel type.  Only *visible* boards are listed; dimensions and thickness are
    measured from the real 3D geometry; identical pieces within a cabinet
    (same spec, size, decor and banding pattern) collapse into one row with a
    quantity."""
    from . import utils

    rows = []
    for prefix in xls.get_prefixes(design):
        flags = xls.cabinet_flags(design, prefix)
        cabinet = prefix.rstrip("_")

        boards = []
        wrapper = next((o for o in design.rootComponent.occurrences
                        if o.component.name == cabinet), None)
        if wrapper is not None:
            xls.collect_boards(wrapper, boards, prefix)
        else:
            for occ in design.rootComponent.occurrences:
                xls.collect_boards(occ, boards, prefix)
        holder = wrapper if wrapper is not None else design.rootComponent

        decor_cache = {}

        def _decor(spec_name):
            if spec_name not in decor_cache:
                decor_cache[spec_name] = utils.effective_decor(
                    design, prefix, spec_name, holder, flags
                )
            return decor_cache[spec_name]

        banded_cache = {}

        def _banded(spec_name):
            if spec_name not in banded_cache:
                banded_cache[spec_name] = utils.effective_banding(
                    design, prefix, spec_name, holder, flags
                )
            return banded_cache[spec_name]

        groups = {}
        for spec_name, body in boards:
            measured = xls.board_geometry_dims(body)
            if measured is None:
                futil.log(f"Iverpan cut list: could not measure a '{prefix}{spec_name}' body",
                          adsk.core.LogLevels.WarningLogLevel)
                continue
            duljina_cm, sirina_cm, debljina_cm, world_dims = measured
            decor = _decor(spec_name)
            banded = _banded(spec_name)
            long_decors, short_decors = base_design.banding_decors_for_orientations(
                spec_name, banded, world_dims
            )
            duljina_mm = round(duljina_cm * 10)
            sirina_mm = round(sirina_cm * 10)
            debljina_mm = round(debljina_cm * 10, 1)
            key = (decor, debljina_mm, spec_name, duljina_mm, sirina_mm,
                   tuple(long_decors), tuple(short_decors))
            if key in groups:
                groups[key]["qty"] += 1
            else:
                groups[key] = {
                    "spec": spec_name, "decor": decor, "debljina": debljina_mm,
                    "duljina": duljina_mm, "sirina": sirina_mm,
                    "long_decors": long_decors, "short_decors": short_decors,
                    "qty": 1,
                }

        def _order(name):
            return xls.row_order.index(name) if name in xls.row_order else len(xls.row_order)

        for key in sorted(groups, key=lambda k: (_order(k[2]), -k[3], -k[4])):
            g = groups[key]
            label = xls.napomena_by_spec.get(g["spec"], g["spec"])
            rows.append({
                "materijal": g["decor"],
                "debljina": g["debljina"],
                "duljina": g["duljina"],
                "sirina": g["sirina"],
                "kom": g["qty"],
                "long_decors": g["long_decors"],
                "short_decors": g["short_decors"],
                "napomena": f"{cabinet} {label}" if cabinet else label,
            })
    return rows


# ---------------------------------------------------------------------------
# Excel writing.
# ---------------------------------------------------------------------------
def _template_path():
    # iverpan_export.py -> commandDialog -> commands -> repo root
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(root, "Iverpan tablica za narudzbu.xlsx")


def _default_output_path(design):
    """Where to drop the exported copy: the user's saved export folder (see
    `excel_export.get_export_folder`), else the Desktop, else the home
    directory, named after the active document.  One file for the whole
    design (Iverpan rows carry their own material/thickness)."""
    app = adsk.core.Application.get()
    doc_name = re.sub(r'[\\/:*?"<>|]+', "_", app.activeDocument.name or "ormar").strip() or "ormar"
    folder = xls._default_export_folder()
    return os.path.join(folder, f"{doc_name}-iverpan-krojna-lista.xlsx")


def _band_cell(decors, index):
    return decors[index] if index < len(decors) else None


def _write_to_excel(rows, template_path, output_path):
    """Load the template, clear the old data band, write the rows, and save a
    copy to output_path.  The template file is never modified."""
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    ws = wb[_SHEET_NAME]

    for r in range(_FIRST_DATA_ROW, _LAST_TEMPLATE_ROW + 1):
        for c in range(_COL_MATERIJAL, _LAST_CLEAR_COL + 1):
            ws.cell(row=r, column=c).value = None

    for i, row in enumerate(rows):
        r = _FIRST_DATA_ROW + i
        ws.cell(row=r, column=_COL_RB).value = i + 1  # numbers rows past the band too
        ws.cell(row=r, column=_COL_MATERIJAL).value = row["materijal"]
        ws.cell(row=r, column=_COL_DEBLJINA).value = row["debljina"]
        ws.cell(row=r, column=_COL_MJERA1).value = row["duljina"]
        ws.cell(row=r, column=_COL_MJERA2).value = row["sirina"]
        ws.cell(row=r, column=_COL_KOM).value = row["kom"]
        ws.cell(row=r, column=_COL_KANT_1A).value = _band_cell(row["long_decors"], 0)
        ws.cell(row=r, column=_COL_KANT_1B).value = _band_cell(row["long_decors"], 1)
        ws.cell(row=r, column=_COL_KANT_2A).value = _band_cell(row["short_decors"], 0)
        ws.cell(row=r, column=_COL_KANT_2B).value = _band_cell(row["short_decors"], 1)
        ws.cell(row=r, column=_COL_NAPOMENA).value = row["napomena"]

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Entry point (called from the dialog's export button).
# ---------------------------------------------------------------------------
def export_cut_list():
    """Export the active design's cut list to one .xlsx copy of the Iverpan
    order form.  Material and edge-band cells hold our own decor names as
    placeholders -- see the module docstring -- the user must replace them
    with real Iverpan codes before sending the order.

    Returns the output path on success, or ``None`` if the design has no
    cabinets (nothing to export).  If the target filename already exists, the
    user is asked once (via `excel_export.resolve_output_paths`) whether to
    overwrite it or save it under a new ``' (N)'``-suffixed name.  Raises on
    I/O / openpyxl errors so the caller can surface them.
    """
    app = adsk.core.Application.get()
    design = adsk.fusion.Design.cast(app.activeProduct)
    if design is None:
        return None

    rows = _gather_rows(design)
    if not rows:
        return None

    output_path, = xls.resolve_output_paths([_default_output_path(design)])
    _write_to_excel(rows, _template_path(), output_path)
    futil.log(f"Iverpan cut list exported: {len(rows)} rows -> {output_path}")
    return output_path
